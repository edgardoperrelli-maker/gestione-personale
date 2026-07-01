#!/usr/bin/env python3
"""Rigenera il template Excel della Produzione economica (Dashboard + grafici NATIVI).

Uso:  python scripts/build-dashboard-template.py   (dalla radice del repo)

Produce:
- public/templates/Produzione-Economica-Dashboard.xlsx   (sorgente editabile a mano in Excel)
- lib/produzione/templateDashboard.json                  (base64 usato a runtime dall'endpoint export)

L'endpoint /api/admin/acea/produzione/export legge il JSON, inietta i dati nelle celle del foglio
"Dati"/"Dettaglio"/"Audit" via jszip (lib/produzione/excelInject.ts) e i grafici si aggiornano da soli.
NB: i grafici leggono SOLO il blocco voci del foglio "Dati" (categorie fisse EL/ES/ERC/ERA), quindi
non dipendono da righe a lunghezza variabile. Se cambi il layout qui, aggiorna anche mappaCelleProduzione.
"""
import base64, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference

NAVY, EUR, ARIAL = "0F2749", '#,##0.00 "€"', "Arial"
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

wb = Workbook()
hdr = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor=NAVY)

# ---- DATI (i grafici/KPI leggono il blocco voci; l'app inietta qui i valori) ----
dati = wb.active
dati.title = "Dati"
for c, t in (("A1", "Voce"), ("B1", "Ordini"), ("C1", "Produzione"), ("D1", "SAL")):
    dati[c] = t; dati[c].font = hdr; dati[c].fill = hfill
for i, v in enumerate(["EL", "ES", "ERC", "ERA", "NON_RISOLTA"]):
    r = 2 + i
    dati[f"A{r}"] = v; dati[f"A{r}"].font = Font(name=ARIAL)
    for col in ("B", "C", "D"):
        dati[f"{col}{r}"] = 0
        if col in ("C", "D"):
            dati[f"{col}{r}"].number_format = EUR
dati["A9"] = "Periodo dal"; dati["A9"].font = Font(name=ARIAL, bold=True); dati["B9"] = "-"
dati["A10"] = "al"; dati["A10"].font = Font(name=ARIAL, bold=True); dati["B10"] = "-"
dati.column_dimensions["A"].width = 16
for col in ("B", "C", "D"):
    dati.column_dimensions[col].width = 14

# ---- DASHBOARD ----
dash = wb.create_sheet("Dashboard")
dash.sheet_view.showGridLines = False
dash["A1"] = "Produzione economica — ACEA"
dash["A1"].font = Font(name=ARIAL, bold=True, size=16, color=NAVY); dash.merge_cells("A1:F1")
dash["A2"] = '="Periodo "&Dati!B9&"  →  "&Dati!B10'
dash["A2"].font = Font(name=ARIAL, size=10, color="64748B")
for lc, lt, vc, vf, nf in [
    ("A4", "Produzione", "A5", "=SUM(Dati!C2:C6)", EUR),
    ("B4", "SAL", "B5", "=SUM(Dati!D2:D6)", EUR),
    ("C4", "Scarto", "C5", "=SUM(Dati!C2:C6)-SUM(Dati!D2:D6)", EUR),
    ("D4", "Ordini", "D5", "=SUM(Dati!B2:B6)", "#,##0"),
]:
    dash[lc] = lt; dash[lc].font = Font(name=ARIAL, size=9, color="64748B")
    dash[vc] = vf; dash[vc].font = Font(name=ARIAL, bold=True, size=14, color=NAVY)
    dash[vc].number_format = nf; dash[vc].fill = PatternFill("solid", fgColor="F1F5F9")
for col in ("A", "B", "C", "D"):
    dash.column_dimensions[col].width = 20
cats = Reference(dati, min_col=1, min_row=2, max_row=6)
c1 = BarChart(); c1.type = "col"; c1.title = "Produzione € per voce"; c1.legend = None
c1.y_axis.numFmt = "#,##0"
c1.add_data(Reference(dati, min_col=3, min_row=1, max_row=6), titles_from_data=True)
c1.set_categories(cats); c1.height = 7.5; c1.width = 15
dash.add_chart(c1, "A8")
c2 = BarChart(); c2.type = "col"; c2.grouping = "clustered"; c2.title = "Produzione vs SAL per voce"
c2.y_axis.numFmt = "#,##0"
c2.add_data(Reference(dati, min_col=3, max_col=4, min_row=1, max_row=6), titles_from_data=True)
c2.set_categories(cats); c2.height = 7.5; c2.width = 15
dash.add_chart(c2, "A24")

# ---- DETTAGLIO (tabelle injected, no grafici) ----
det = wb.create_sheet("Dettaglio")
def block(start_col, titolo, rows):
    for j, t in enumerate([titolo, "Ordini", "Produzione"]):
        cell = det.cell(row=1, column=start_col + j, value=t); cell.font = hdr; cell.fill = hfill
    for r in range(2, 2 + rows):
        det.cell(row=r, column=start_col, value="")
        det.cell(row=r, column=start_col + 1, value=0)
        det.cell(row=r, column=start_col + 2, value=0).number_format = EUR
    det.column_dimensions[chr(64 + start_col)].width = 26
block(1, "Operatore", 15)
block(5, "Territorio", 15)
det.cell(row=1, column=9, value="Giorno").font = hdr; det.cell(row=1, column=9).fill = hfill
det.cell(row=1, column=10, value="Produzione").font = hdr; det.cell(row=1, column=10).fill = hfill
for r in range(2, 33):
    det.cell(row=r, column=9, value="")
    det.cell(row=r, column=10, value=0).number_format = EUR
det.column_dimensions["I"].width = 14; det.column_dimensions["J"].width = 14
# blocco ATTIVITÀ (L/M/N), fino a 40 righe
for j, t in enumerate(["Attività", "Ordini", "Produzione"]):
    c = det.cell(row=1, column=12 + j, value=t); c.font = hdr; c.fill = hfill
for r in range(2, 42):
    det.cell(row=r, column=12, value="")
    det.cell(row=r, column=13, value=0)
    det.cell(row=r, column=14, value=0).number_format = EUR
det.column_dimensions["L"].width = 32; det.column_dimensions["N"].width = 14

# ---- AUDIT (tabella injected) ----
au = wb.create_sheet("Audit")
au.cell(row=1, column=1, value="ODL").font = hdr; au.cell(row=1, column=1).fill = hfill
au.cell(row=1, column=2, value="Discrepanza").font = hdr; au.cell(row=1, column=2).fill = hfill
for r in range(2, 202):
    au.cell(row=r, column=1, value=""); au.cell(row=r, column=2, value="")
au.column_dimensions["A"].width = 22; au.column_dimensions["B"].width = 52

wb.move_sheet("Dashboard", -wb.sheetnames.index("Dashboard"))  # Dashboard primo

xlsx_path = os.path.join(ROOT, "public", "templates", "Produzione-Economica-Dashboard.xlsx")
os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
wb.save(xlsx_path)
b = open(xlsx_path, "rb").read()
json.dump({"b64": base64.b64encode(b).decode()},
          open(os.path.join(ROOT, "lib", "produzione", "templateDashboard.json"), "w"))
print("OK:", xlsx_path, f"({len(b)} bytes) + templateDashboard.json")
